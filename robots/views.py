from datetime import timedelta, date
from io import BytesIO
from typing import Any

from django.db.models import Count
from django.db.models.query import ValuesQuerySet
from django.http import HttpRequest, FileResponse
from django.utils import timezone
from django.views.generic import TemplateView
from openpyxl import Workbook

from .models import Robot


class RobotReportTemplateView(TemplateView):
    """
    Класс представления для создания Excel-файла со сводкой
    по суммарным показателям производства роботов за последнюю неделю.
    """
    def get_robot_report(self) -> Workbook:
        """Создает отчет на основании даных из таблицы Robots."""
        today: date = timezone.now().date()
        report_data: ValuesQuerySet[Robot, dict[str, Any]] = (
            Robot.objects.filter(
                    created__gte=today - timedelta(days=7),
                    created__lt=today
                )
            .values('model', 'version')
            .annotate(version_count=Count('version'))
            .order_by('-version_count')
        )
        wb: Workbook = Workbook()
        headers: list[str] = [
            'Модель', 'Версия', 'Количество за неделю'
        ]
        for robot in report_data:
            model: str = robot['model']
            if model not in wb.sheetnames:
                wb.create_sheet(model)
                wb[model].append(headers)
            wb[model].append(
                [model,
                 robot['version'],
                 robot['version_count']]
            )
        wb.remove_sheet(wb['Sheet'])
        return wb

    def get(
            self, request: HttpRequest, *args: Any, **kwargs: Any
    ) -> FileResponse:
        """
        Обрабатывает GET-запрос и возвращает
        файловый ответ с отчетом в формате Excel.
        """
        wb: Workbook = self.get_robot_report()
        buffer: BytesIO = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return FileResponse(
            buffer,
            filename=f'report_{timezone.now().date()}.xlsx'
        )
